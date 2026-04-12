#!/usr/bin/env python3
"""
Excel 报告生成器
================
将每日 CSV 数据转换为格式化的 .xlsx Excel 文件，
输出到 docs/data/ 目录供网页端直接下载。

用法：
    python3 generate_excel.py                          # 默认
    python3 generate_excel.py --output-dir ./output --site-dir ./docs
"""

import csv
import os
import sys
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠️  openpyxl 未安装，正在安装...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

import re

# 清除 Excel 不支持的非法字符（控制字符等）
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def clean_for_excel(val):
    """清除 Excel 不支持的非法字符"""
    if not isinstance(val, str):
        return val
    return ILLEGAL_CHARS_RE.sub('', val)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_OUTPUT_DIR = os.path.join(BASE_DIR, "output")
DEFAULT_SITE_DIR = os.path.join(BASE_DIR, "docs")


def read_csv_rows(csv_path):
    """读取 CSV 为字典列表"""
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for row in reader:
            rows.append(dict(row))
    return headers, rows


def find_csvs(output_dir):
    """查找所有日期的 CSV 文件"""
    import glob
    result = {}
    for f in glob.glob(os.path.join(output_dir, "每日热点汇总表_*.csv")):
        basename = os.path.basename(f)
        date_str = basename.replace("每日热点汇总表_", "").replace(".csv", "")[:10]
        if len(date_str) == 10 and date_str[4] == "-":
            if date_str not in result or f > result[date_str]:
                result[date_str] = f
    archive_dir = os.path.join(output_dir, "archive")
    if os.path.isdir(archive_dir):
        for d in os.listdir(archive_dir):
            date_dir = os.path.join(archive_dir, d)
            if not os.path.isdir(date_dir):
                continue
            for f in glob.glob(os.path.join(date_dir, "每日热点汇总表_*.csv")):
                date_str = d[:10]
                if date_str not in result or f > result[date_str]:
                    result[date_str] = f
    return result


# 样式定义
HEADER_FILL = PatternFill(start_color="1e1b4b", end_color="1e1b4b", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=9)
P0_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
P1_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# 要输出的列（精选关键列，排除过长的内容摘要）
EXPORT_COLUMNS = [
    "采集日期", "栏目", "来源Key", "标题", "链接", "中文简介",
    "是否AI相关", "是否云行业", "综合评分", "热度评分", "时效评分",
    "与腾讯云结合度", "产品标签", "竞品标签", "友商标签", "技术标签",
    "官方号主推产品", "发布优先级", "腾讯云结合点", "官方号写作角度",
    "社媒结论", "开发者/作者", "开发者链接", "开发者邮箱",
]

# 列宽配置
COL_WIDTHS = {
    "采集日期": 12, "栏目": 10, "来源Key": 12, "标题": 40, "链接": 30,
    "中文简介": 35, "是否AI相关": 8, "是否云行业": 8,
    "综合评分": 8, "热度评分": 8, "时效评分": 8, "与腾讯云结合度": 10,
    "产品标签": 15, "竞品标签": 15, "友商标签": 15, "技术标签": 18,
    "官方号主推产品": 14, "发布优先级": 10,
    "腾讯云结合点": 40, "官方号写作角度": 35,
    "社媒结论": 35, "开发者/作者": 14, "开发者链接": 25, "开发者邮箱": 20,
}


def create_daily_excel(date_str, csv_path, output_path):
    """为单日数据生成 Excel"""
    headers, rows = read_csv_rows(csv_path)
    if not rows:
        return False

    wb = openpyxl.Workbook()

    # === Sheet 1: 全量数据 ===
    ws = wb.active
    ws.title = "全量数据"

    # 写表头
    cols = [c for c in EXPORT_COLUMNS if c in headers or c == "采集日期"]
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 15)

    # 写数据
    for ri, row in enumerate(rows, 2):
        for ci, col_name in enumerate(cols, 1):
            val = row.get(col_name, "")
            if col_name == "采集日期" and not val:
                val = date_str
            # 截断超长文本
            if isinstance(val, str) and len(val) > 500:
                val = val[:497] + "..."
            cell = ws.cell(row=ri, column=ci, value=clean_for_excel(val))
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        # P0/P1 行高亮
        priority = row.get("发布优先级", "") or ""
        if priority.startswith("P0"):
            for ci in range(1, len(cols) + 1):
                ws.cell(row=ri, column=ci).fill = P0_FILL
        elif priority.startswith("P1"):
            for ci in range(1, len(cols) + 1):
                ws.cell(row=ri, column=ci).fill = P1_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # === Sheet 2: 精选话题 ===
    ws2 = wb.create_sheet("精选话题")
    selected = [r for r in rows
                if ((r.get("发布优先级") or "").startswith("P0") or (r.get("发布优先级") or "").startswith("P1"))
                and len((r.get("腾讯云结合点") or "").strip()) > 5
                and (r.get("腾讯云结合点") or "").strip() != "无"]

    # 精选表头
    selected_cols = ["标题", "来源Key", "综合评分", "与腾讯云结合度", "发布优先级",
                     "产品标签", "腾讯云结合点", "官方号写作角度", "社媒结论", "链接"]
    for ci, col_name in enumerate(selected_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 18)

    for ri, row in enumerate(selected, 2):
        for ci, col_name in enumerate(selected_cols, 1):
            val = row.get(col_name, "")
            if isinstance(val, str) and len(val) > 500:
                val = val[:497] + "..."
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

    ws2.auto_filter.ref = ws2.dimensions if selected else "A1:J1"
    ws2.freeze_panes = "A2"

    wb.save(output_path)
    return True


def create_master_excel(output_dir, output_path):
    """生成历史总表 Excel"""
    master_csv = os.path.join(output_dir, "热点追踪历史总表.csv")
    if not os.path.exists(master_csv):
        print("⚠️  历史总表不存在，跳过")
        return False

    headers, rows = read_csv_rows(master_csv)
    if not rows:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "历史总表"

    cols = [c for c in EXPORT_COLUMNS if c in headers]
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 15)

    for ri, row in enumerate(rows, 2):
        for ci, col_name in enumerate(cols, 1):
            val = row.get(col_name, "")
            if isinstance(val, str) and len(val) > 500:
                val = val[:497] + "..."
            cell = ws.cell(row=ri, column=ci, value=clean_for_excel(val))
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        priority = row.get("发布优先级", "") or ""
        if priority.startswith("P0"):
            for ci in range(1, len(cols) + 1):
                ws.cell(row=ri, column=ci).fill = P0_FILL
        elif priority.startswith("P1"):
            for ci in range(1, len(cols) + 1):
                ws.cell(row=ri, column=ci).fill = P1_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return True


def generate_excels(output_dir, site_dir):
    """生成所有 Excel 文件"""
    data_dir = os.path.join(site_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    csvs = find_csvs(output_dir)
    if not csvs:
        print("⚠️  未找到 CSV 文件")
        return

    dates = sorted(csvs.keys(), reverse=True)
    print(f"📊 生成 Excel 报告...")

    # 最新一天的日报 Excel
    latest = dates[0]
    daily_path = os.path.join(data_dir, f"hotspot-daily-{latest}.xlsx")
    if create_daily_excel(latest, csvs[latest], daily_path):
        print(f"  ✅ 日报 Excel: {daily_path}")

    # 历史总表 Excel
    master_path = os.path.join(data_dir, "hotspot-master.xlsx")
    if create_master_excel(output_dir, master_path):
        print(f"  ✅ 历史总表 Excel: {master_path}")

    print(f"🎉 Excel 生成完成！")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成 Excel 报告")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--site-dir", default=DEFAULT_SITE_DIR)
    args = parser.parse_args()
    generate_excels(args.output_dir, args.site_dir)
